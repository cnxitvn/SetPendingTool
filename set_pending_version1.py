#process excel file
import win32com.client
import time
from openpyxl import load_workbook
import psutil
import win32com.client
import os
def checknosleep():
    import psutil
    import subprocess

    nosleep_path = r"C:\Softs\NoSleep.exe"
    found = False

# Check all running processes
    for proc in psutil.process_iter(['pid', 'name', 'username']):
        try:
            if proc.info['name'] and "NoSleep" in proc.info['name'] and \
            proc.info['username'] and "CONCENTRIX\\congthang.van" in proc.info['username']:
                print(f"Process found: {proc.info['name']} (PID: {proc.info['pid']}, User: {proc.info['username']})")
                found = True
                break
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass

# If not found, run NoSleep.exe
    if not found:
        print("NoSleep.exe not found. Launching...")
        subprocess.Popen(nosleep_path)

checknosleep()

def send_troubleshooting_email(ticketnumber, requester_email):
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.application import MIMEApplication
    import os

    # Zimbra SMTP configuration
    smtp_server = "zimbra-smtp.concentrix.com"
    smtp_port = 465
    user_email = "itvn_noreply@concentrix.com"
    user_password = "input_your_password_here"  # Replace with the actual password or use environment variables for security
    cc_email = "input_email_you_want_to_cc"
    # Email content
    recipient_emails = [requester_email,
                        cc_email
    ]
    subject = f"""Incident Acknowledgement || {ticketnumber}"""

    # HTML body
    html_body = f"""
<p style="margin: 0cm; font-family: Aptos, sans-serif;"><span style="font-size: 12pt; font-family: 'times new roman', times, serif;">Dear {requester_email},</span></p>
<p style="margin: 0cm; font-family: Aptos, sans-serif;">&nbsp;</p>
<p style="margin: 0cm; font-size: 12pt; font-family: Aptos, sans-serif;"><span style="font-size: 12pt; font-family: 'times new roman', times, serif;">Thank you for patience.&nbsp;<span data-teams="true">Your request is being processed. We will inform you should we need any additional information.</span></span></p>
<p style="margin: 0cm; font-family: Aptos, sans-serif;">&nbsp;</p>
<p style="margin: 0cm; font-family: Aptos, sans-serif;"><span style="font-family: 'times new roman', times, serif; font-size: 16px;"><strong>Thanks,&nbsp;</strong></span></p>
<p style="line-height: 12.65pt; background: white; margin: 0cm; font-size: 12pt; font-family: Aptos, sans-serif;"><span style="font-family: 'times new roman', times, serif; font-size: 16px;"><strong><span style="color: #002451;">Click below to be directed to specific ticket types:</span></strong></span></p>
<p style="line-height: 12.65pt; background: white; margin: 0cm; font-size: 12pt; font-family: Aptos, sans-serif;"><span style="font-family: 'times new roman', times, serif; font-size: 16px;"><u><span style="color: blue;"><a title="https://solvnow-dwp.onbmc.com/dwp/app/#/itemprofile/183" href="https://nam10.safelinks.protection.outlook.com/?url=https%3A%2F%2Fsolvnow-dwp.onbmc.com%2Fdwp%2Fapp%2F%23%2Fitemprofile%2F183&amp;data=05%7C02%7Ccongthang.van%40concentrix.com%7C99ecb31eb39f4c46d3cf08ddae6f2dcc%7C599e51d62f8c43478e591f795a51a98c%7C0%7C0%7C638858514716686852%7CUnknown%7CTWFpbGZsb3d8eyJFbXB0eU1hcGkiOnRydWUsIlYiOiIwLjAuMDAwMCIsIlAiOiJXaW4zMiIsIkFOIjoiTWFpbCIsIldUIjoyfQ%3D%3D%7C0%7C%7C%7C&amp;sdata=Kt4FWRAlEcX2iSNZmnrN4Pc5igxCuDM2HGvYkvdp7e4%3D&amp;reserved=0" data-linkindex="0" data-auth="NotApplicable">Office 365</a>&nbsp;| <a title="https://solvnow-dwp.onbmc.com/dwp/app/#/itemprofile/172" href="https://nam10.safelinks.protection.outlook.com/?url=https%3A%2F%2Fsolvnow-dwp.onbmc.com%2Fdwp%2Fapp%2F%23%2Fitemprofile%2F172&amp;data=05%7C02%7Ccongthang.van%40concentrix.com%7C99ecb31eb39f4c46d3cf08ddae6f2dcc%7C599e51d62f8c43478e591f795a51a98c%7C0%7C0%7C638858514716713999%7CUnknown%7CTWFpbGZsb3d8eyJFbXB0eU1hcGkiOnRydWUsIlYiOiIwLjAuMDAwMCIsIlAiOiJXaW4zMiIsIkFOIjoiTWFpbCIsIldUIjoyfQ%3D%3D%7C0%7C%7C%7C&amp;sdata=lO5nzz%2FQZksX80IDiVCFphio7I10GhL%2BwViL1a5afgQ%3D&amp;reserved=0" data-linkindex="1" data-auth="NotApplicable">Zimbra</a></span></u><span style="color: blue;"><a title="https://solvnow-dwp.onbmc.com/dwp/app/#/itemprofile/172" href="https://nam10.safelinks.protection.outlook.com/?url=https%3A%2F%2Fsolvnow-dwp.onbmc.com%2Fdwp%2Fapp%2F%23%2Fitemprofile%2F172&amp;data=05%7C02%7Ccongthang.van%40concentrix.com%7C99ecb31eb39f4c46d3cf08ddae6f2dcc%7C599e51d62f8c43478e591f795a51a98c%7C0%7C0%7C638858514716729189%7CUnknown%7CTWFpbGZsb3d8eyJFbXB0eU1hcGkiOnRydWUsIlYiOiIwLjAuMDAwMCIsIlAiOiJXaW4zMiIsIkFOIjoiTWFpbCIsIldUIjoyfQ%3D%3D%7C0%7C%7C%7C&amp;sdata=tYfifZtcfQepgXf8%2FswTMOjhzkyto%2BRnO82D72h1iRg%3D&amp;reserved=0" data-linkindex="2" data-auth="NotApplicable">&nbsp;</a><u>| <a title="https://solvnow-dwp.onbmc.com/dwp/app/#/itemprofile/142" href="https://nam10.safelinks.protection.outlook.com/?url=https%3A%2F%2Fsolvnow-dwp.onbmc.com%2Fdwp%2Fapp%2F%23%2Fitemprofile%2F142&amp;data=05%7C02%7Ccongthang.van%40concentrix.com%7C99ecb31eb39f4c46d3cf08ddae6f2dcc%7C599e51d62f8c43478e591f795a51a98c%7C0%7C0%7C638858514716743702%7CUnknown%7CTWFpbGZsb3d8eyJFbXB0eU1hcGkiOnRydWUsIlYiOiIwLjAuMDAwMCIsIlAiOiJXaW4zMiIsIkFOIjoiTWFpbCIsIldUIjoyfQ%3D%3D%7C0%7C%7C%7C&amp;sdata=no1nFVLrfOK1rmznSfYD%2BruCZ4SPnZ3QBFapnp3zmXo%3D&amp;reserved=0" data-linkindex="3" data-auth="NotApplicable">Computer &amp; Accessories</a>&nbsp;| <a title="https://solvnow-dwp.onbmc.com/dwp/app/#/itemprofile/138" href="https://nam10.safelinks.protection.outlook.com/?url=https%3A%2F%2Fsolvnow-dwp.onbmc.com%2Fdwp%2Fapp%2F%23%2Fitemprofile%2F138&amp;data=05%7C02%7Ccongthang.van%40concentrix.com%7C99ecb31eb39f4c46d3cf08ddae6f2dcc%7C599e51d62f8c43478e591f795a51a98c%7C0%7C0%7C638858514716757651%7CUnknown%7CTWFpbGZsb3d8eyJFbXB0eU1hcGkiOnRydWUsIlYiOiIwLjAuMDAwMCIsIlAiOiJXaW4zMiIsIkFOIjoiTWFpbCIsIldUIjoyfQ%3D%3D%7C0%7C%7C%7C&amp;sdata=AKI8RCGoXgJ6Wrgi7dELUOUhaioxH53zw1Qd7I0hbL4%3D&amp;reserved=0" data-linkindex="4" data-auth="NotApplicable">&nbsp;Application Product</a>&nbsp;| <a title="https://solvnow-dwp.onbmc.com/dwp/app/#/itemprofile/126" href="https://nam10.safelinks.protection.outlook.com/?url=https%3A%2F%2Fsolvnow-dwp.onbmc.com%2Fdwp%2Fapp%2F%23%2Fitemprofile%2F126&amp;data=05%7C02%7Ccongthang.van%40concentrix.com%7C99ecb31eb39f4c46d3cf08ddae6f2dcc%7C599e51d62f8c43478e591f795a51a98c%7C0%7C0%7C638858514716770423%7CUnknown%7CTWFpbGZsb3d8eyJFbXB0eU1hcGkiOnRydWUsIlYiOiIwLjAuMDAwMCIsIlAiOiJXaW4zMiIsIkFOIjoiTWFpbCIsIldUIjoyfQ%3D%3D%7C0%7C%7C%7C&amp;sdata=oVbm7kdmM7lfq1ZZmMeseUi3Jkq9kCa7iqJFjUiQGEg%3D&amp;reserved=0" data-linkindex="5" data-auth="NotApplicable">Password Reset</a>&nbsp;| &nbsp;<a title="https://solvnow-dwp.onbmc.com/dwp/app/#/page/mtaslejl" href="https://nam10.safelinks.protection.outlook.com/?url=https%3A%2F%2Fsolvnow-dwp.onbmc.com%2Fdwp%2Fapp%2F%23%2Fpage%2Fmtaslejl&amp;data=05%7C02%7Ccongthang.van%40concentrix.com%7C99ecb31eb39f4c46d3cf08ddae6f2dcc%7C599e51d62f8c43478e591f795a51a98c%7C0%7C0%7C638858514716783836%7CUnknown%7CTWFpbGZsb3d8eyJFbXB0eU1hcGkiOnRydWUsIlYiOiIwLjAuMDAwMCIsIlAiOiJXaW4zMiIsIkFOIjoiTWFpbCIsIldUIjoyfQ%3D%3D%7C0%7C%7C%7C&amp;sdata=fyIoehTFirL%2B76HK1bCgr3UDYAM1LpW%2Btr7Nww6k7%2Bc%3D&amp;reserved=0" data-linkindex="6" data-auth="NotApplicable">SolvNow</a></u></span></span></p>
<p style="margin: 0cm; font-size: 12pt; font-family: Aptos, sans-serif;"><span style="font-family: 'times new roman', times, serif; font-size: 16px;"><strong><em>Please always keep (<a href="mailto:VN_DesksideSupport_Local_IT@concentrix.com"><span style="color: #467886;">VN_DesksideSupport_Local_IT@concentrix.com</span></a>) in loop for faster response. We are available to support you.</em></strong></span></p>
"""
    #create the email message
    msg = MIMEMultipart()
    msg['From'] = user_email
    msg['To'] = ", ".join(recipient_emails)
    msg['Subject'] = subject
    msg.attach(MIMEText(html_body, 'html'))
    try:
        server = smtplib.SMTP_SSL(smtp_server, smtp_port)
        server.login(user_email, user_password)
        server.sendmail(user_email, recipient_emails, msg.as_string())
        server.quit()
        print("Email sent successfully.")
    except Exception as e:
        print(f"Error: {e}")


def getticket_number(ticket):
    excel_file_done = r"C:\Users\congthang.van\OneDrive - Concentrix Corporation\process_pending_ticket\set_pending_done.xlsx" #input your path to your excel file
    # excel_file_done = r"C:\Users\ManageITVN\OneDrive - Concentrix Corporation\process_pending_ticket\set_pending_done.xlsx"
    wb_pending = load_workbook(excel_file_done)
    sheet = wb_pending.active
    last_row = sheet.max_row + 1
    sheet.cell(row=last_row, column=1).value = ticket
    last_row += 1
    wb_pending.save(excel_file_done)
    wb_pending.close()
set_pending_file = r"C:\Users\congthang.van\OneDrive - Concentrix Corporation\process_pending_ticket\set_pending_new.xlsx"   #input your
set_pending_file_temp = set_pending_file + ".temp_check"
def is_file_locked(file_path, temp_path):
    try:
        os.rename(file_path, temp_path)
        os.rename(temp_path, file_path)
        return False  # File is not locked
    except PermissionError:
        return True  # File is locked (likely open in Excel)
def process_excel_file(file_path):
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = True  
    wb = xl.Workbooks.Open(set_pending_file)
    # wb=xl.Workbooks.Open(r"C:\Users\ManageITVN\OneDrive - Concentrix Corporation\process_pending_ticket\set_pending_new.xlsx")
    wb.RefreshAll()
    time.sleep(18)
    wb.Save()
    wb.Close()
    xl.Quit()

    del wb, xl
    time.sleep(5)

def kill_excel_processes():
    for proc in psutil.process_iter(['name']):
        if proc.info['name'] and 'EXCEL.EXE' in proc.info['name'].upper():
            proc.kill()
            print("Killed:", proc.info['name'])

if is_file_locked(set_pending_file, set_pending_file_temp):
    print(f"{set_pending_file} is open. Killing Excel processes...")
    kill_excel_processes()
    time.sleep(5)  # Wait a moment to ensure processes are killed
    process_excel_file(set_pending_file)
else:
    print(f"{set_pending_file} is not open.")
    process_excel_file(set_pending_file)



from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
chrome_options = Options()
chrome_options.add_argument(r'--user-data-dir=C:\Users\congthang.van\AppData\Local\Google\Chrome\User Data\Default')  # Path to your Chrome profile
# Optionally, specify your profile directory if you use multiple profiles:
# chrome_options.add_argument(r'--profile-directory=Default')  # Or 'Profile 1', etc.


baseURL = "https://solvnow-smartit.onbmc.com/smartit/app/#/search/"

# Path to your Excel file
# excel_file = r"C:\Users\ManageITVN\OneDrive - Concentrix Corporation\process_pending_ticket\set_pending_new.xlsx"
excel_file = r"C:\Users\congthang.van\OneDrive - Concentrix Corporation\process_pending_ticket\set_pending_new.xlsx" 
# C:\Users\congthang.van\OneDrive - Concentrix Corporation\process_pending_ticket\set_pending_new.xlsx   
ticket_success = []
# Read phrases from Excel
import pandas as pd


df = pd.read_excel(excel_file,sheet_name="data") 
dftickets = df[df['check1']==0]# Assuming column name is "ticket"
# dftickets
tickets = dftickets['Data.ticket'].tolist()  # Adjust column name if necessary
tickets = list(set(tickets))  # Remove duplicates
print(f"Tickets to process: {tickets}")
for ticket in tickets:
        driver = webdriver.Chrome(options=chrome_options)
        driver.get(baseURL+ticket)
        print(f"Processing ticket: {ticket}")
        try:

            time.sleep(20)
            onehub = 'OneHub'   #please change to your desired group
            qtsc = 'HoChiMinh' #please change to your desired group
            flemington = 'Flemington' #please change to your desired group
            techvalley = 'TechValley' #please change to your desired group
            assign_group = driver.find_element(By.CLASS_NAME, 'search-item-layout__desc')

            if onehub in assign_group.text:
                engineer = "Hoang Tuan Kiet Nguyen"
            elif qtsc in assign_group.text:
                engineer = "Hoang Lam Cao"
            elif flemington in assign_group.text:
                engineer = "Quoc Khai Bui"
            elif techvalley in assign_group.text:
                engineer = "Hoang (Daniel ) Tran"   
            else:
                engineer = "Hoang Duy Bao Nguyen"
            print(f'''assign engieener to {engineer}''')
        except Exception as e:
            print(f"Failed to find assign engineer: {e} - try to assign default engineer : Hoang Duy Bao Nguyen")
            engineer = "Hoang Duy Bao Nguyen"
        ####Process the target link
        print(f'''assign engieener to {engineer}''')
        try:
            links = driver.find_elements(By.TAG_NAME, "a")
            hrefs = [link.get_attribute("href") for link in links if link.get_attribute("href")]
            
            if hrefs:
                max_href = max(hrefs, key=len)
                print("Link with maximum length:", max_href)
            else:
                print("No href links found.")    
            # Wait for the element by its link text
            driver.get(max_href)
            # pencil_icon = WebDriverWait(driver, 15).until(
            #     EC.element_to_be_clickable((By.XPATH, "//span[contains(@class, 'd-icon-pencil')]"))
            # )
            time.sleep(10)  # Wait for the page to load
            driver.switch_to.frame("pwa-frame")
            time.sleep(15)  # Wait for the iframe to load

            # if flag_send_email == True:
            #     send_troubleshooting_email(ticket, user_email.text)
            # else:
            #     print("No user email found, skipping email notification.")
            element = driver.find_element(By.XPATH, "//div[@id='ar7_data']")
            if element.text == "Assigned":
                            
                try:
                    user_email = driver.find_element(By.ID, "ar1000000048_data")
                    flag_send_email = True
                        # print("ticket is already assigned.")
                except Exception as e:
                    print(f"Skip this ticket {ticket} because it is not user ticket")
                    flag_send_email = False
                if flag_send_email == True:
                    print(user_email.text)
                    send_troubleshooting_email(ticket, user_email.text)
                else:
                    print("No user email found, skipping email notification.")
                edit_buttons = driver.find_element(By.XPATH, "//span[contains(@class, 'd-icon-pencil undefined ng-star-inserted')]")
                edit_buttons.click()
                time.sleep(10)
                
                input_engineer = driver.find_element(By.ID, "ar1000000218")
            
                input_engineer.clear()
                input_engineer.send_keys(engineer)

                time.sleep(3)  # Wait for suggestions to load
                click_engineer = driver.find_element(By.XPATH, "//div[contains(@class, 'textEllipsis') and @data-testmenu]")
                click_engineer.click()
                save_button = driver.find_element(By.XPATH, "//button[@title='Save ticket']")
                save_button.click()
                time.sleep(10)  # Wait for the save action to complete
                edit_buttons = driver.find_element(By.XPATH, "//span[contains(@class, 'd-icon-pencil undefined ng-star-inserted')]")
                edit_buttons.click()
                time.sleep(5)

                status = driver.find_elements(By.CSS_SELECTOR, "div.rx-select__search-button-title")
                print(status[3].text)
            
                if status[3].text == "Assigned":
                    btns = driver.find_elements(By.CSS_SELECTOR, "button.dropdown-toggle")
                    btns[2].click()  # Click the fourth button (index 3)
                    pending_option = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[.//adapt-highlight[text()='Pending']]"))
                    )
                    pending_option.click()
                    #change reason pending
                    time.sleep(3)
                    btns = driver.find_elements(By.CSS_SELECTOR, "button.dropdown-toggle")
                    btns[3].click()  # Click the fourth button (index 3)
                    Client_Action_Required = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[.//adapt-highlight[text()='Client Action Required']]"))
                    )
                    Client_Action_Required.click()
                    time.sleep(3)
                    save_button = driver.find_element(By.XPATH, "//button[@title='Save ticket']")
                    save_button.click()
                    time.sleep(4)
                    getticket_number(ticket)
                    driver.quit()
            
                else:
                    ...
            else:
                print("Element text is not 'Asigned', skipping edit.")
                getticket_number(ticket)
            getticket_number(ticket)
        except Exception as e:
                print(f"Failed to click on the adjust button or select Pending: {e}")
        driver.quit()
        print(f"Ticket {ticket} processed successfully.")